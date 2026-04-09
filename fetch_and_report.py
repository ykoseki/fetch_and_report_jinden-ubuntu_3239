import argparse
import urllib.request
import urllib.parse
import json
import os
import ast
import calendar
from dotenv import load_dotenv
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter

def get_pleasanter_data(base_url, api_key, site_id, year, month):
    last_day = calendar.monthrange(year, month)[1]
    start_date = f"{year}/{month:02d}/01 00:00:00"
    end_date = f"{year}/{month:02d}/{last_day} 23:59:59.997"
    url = f"{base_url}/api/items/{site_id}/get"
    view = {
        "Id": 0,
        "ColumnFilterHash": {
            "DateA": f"[\"{start_date},{end_date}\"]"
        },
        "ColumnSorterHash": {
            "NumL": "desc"
        }
    }
    payload = {"ApiVersion": 1.1, "ApiKey": api_key, "View": view}
    data = json.dumps(payload).encode('utf-8')
    headers = {'Content-Type': 'application/json'}
    try:
        req = urllib.request.Request(url, data=data, headers=headers)
        with urllib.request.urlopen(req) as resp:
            content = resp.read().decode('utf-8')
            res = json.loads(content)
            return res.get('Response', {}).get('Data', [])
    except Exception as e:
        print(f"Error fetching data: {e}")
        return []

def fetch_master(base_url, api_key, path):
    url = f"{base_url}/api/{path}/get"
    payload = {'ApiVersion': 1.1, 'ApiKey': api_key}
    data = json.dumps(payload).encode('utf-8')
    req = urllib.request.Request(url, data=data, headers={'Content-Type': 'application/json'})
    try:
        with urllib.request.urlopen(req) as resp:
            res = json.loads(resp.read().decode('utf-8'))
            return res.get('Response', {}).get('Data', [])
    except Exception as e:
        print(f"Error fetching {path}: {e}")
        return []

def parse_choices(choices_text):
    mapping = {}
    if not choices_text: return mapping
    lines = choices_text.split('\n')
    for line in lines:
        if ',' in line:
            parts = line.split(',')
            mapping[parts[0].strip()] = parts[1].strip()
    return mapping

def get_actual_width(text):
    if not text: return 0
    text = str(text)
    return sum(2 if ord(c) > 0x7F else 1 for c in text)

def shorten_url(long_url):
    """Shorten URL using TinyURL. (is.gd was too strict for local hostnames)"""
    try:
        api_url = f"https://tinyurl.com/api-create.php?url={urllib.parse.quote(long_url)}"
        with urllib.request.urlopen(api_url) as resp:
            return resp.read().decode('utf-8')
    except Exception as e:
        print(f"Warning: URL shortening (TinyURL) failed: {e}")
        return long_url

def generate_pleasanter_viewer_url(base_url, site_id, year, month, min_score=None):
    last_day = calendar.monthrange(year, month)[1]
    start_date = f"{year}/{month:02d}/01 00:00:00"
    end_date = f"{year}/{month:02d}/{last_day} 23:59:59.997"
    
    view = {
        "Id": 0,
        "ColumnFilterHash": {
            "DateA": f"[\"{start_date},{end_date}\"]"
        },
        "ColumnSorterHash": {
            "NumL": "desc"
        }
    }
    if min_score is not None:
        view["ColumnFilterHash"]["NumL"] = f"[\"{min_score},\"]"
        
    view_json = json.dumps(view, separators=(',', ':'))
    long_url = f"{base_url}/items/{site_id}/index?View={view_json}"
    return long_url

def main():
    parser = argparse.ArgumentParser(description="Generate Pleasanter Report with TinyURL Notification")
    parser.add_argument("--year", type=int, default=2026, help="Target Year (YYYY)")
    parser.add_argument("--month", type=int, default=3, help="Target Month (MM)")
    args = parser.parse_args()

    # --- Configuration ---
    # Load environment variables from .env file
    load_dotenv()
    
    base_url = os.getenv("PLEASANTER_BASE_URL")
    site_id = os.getenv("PLEASANTER_SITE_ID")
    api_key = os.getenv("PLEASANTER_API_KEY")

    # Error checking for missing environment variables
    missing_vars = []
    if not base_url: missing_vars.append("PLEASANTER_BASE_URL")
    if not site_id: missing_vars.append("PLEASANTER_SITE_ID")
    if not api_key: missing_vars.append("PLEASANTER_API_KEY")

    if missing_vars:
        print(f"[ERROR] Missing environment variables: {', '.join(missing_vars)}")
        print("Please check your .env file.")
        return

    # --- Configuration (Portable & Slim! [Slim]) ---
    base_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Use environment variables for security! 🛡️
    base_url = os.getenv("PLEASANTER_BASE_URL")
    site_id = os.getenv("PLEASANTER_SITE_ID")
    api_key = os.getenv("PLEASANTER_API_KEY")

    # Files are now OPTIONAL! (Internal defaults used) [DATA]
    excel_output = os.path.join(base_dir, f"Report_{args.year}_{args.month:02d}.xlsx")
    text_output = os.path.join(base_dir, f"Notification_{args.year}_{args.month:02d}.txt")

    # Hardcoded Column Names (Originally from Excel Template) [LIST]
    FINAL_COLUMNS = [
        "ID", "提出日", "氏名", "部署名", "職場", "件名", "区分", 
        "効果金額・率(円・%)", "効果率(%)適用", "改善効果点数", "ヨコテン評価点数", 
        "一次審査日", "一次審査者", "創造力1", "改善力1", "問題解決力1", "合計1", 
        "二次審査日", "二次審査者", "創造力2", "改善力2", "問題解決力2", "合計2", 
        "総合点数", "内容", "受付状況", "ロック", "コメント"
    ]

    # Hardcoded Choices (Fallback dictionary from JSON) [Data]
    CHOICES_FALLBACK = {
        'ClassE': {'10': '第一製造', '11': '品質保証部', '21': '第二製造電顕', '22': '第二製造ME', '31': 'システム営業部', '32': 'サプライ営業部'},
        'Status': {'100': '受付前', '900': '受付済'}
    }

    print(f"\n--- Generating Report and Notification (Slim Edition) ---")
    print(f"[Target] {args.year}年{args.month}月度")
    print(f"[Workspace] {base_dir}")
    print(f"----------------------------------------------------------")

    # 1. Master & Mappings
    print(f"[*] Initializing column mappings...")
    choices_mappings = CHOICES_FALLBACK

    print(f"[*] Fetching user and group masters...")
    users = fetch_master(base_url, api_key, "users")
    groups = fetch_master(base_url, api_key, "groups")
    user_map = {str(u['UserId']): u['Name'] for u in users}
    group_map = {str(g['GroupId']): g['GroupName'] for g in groups}
    print(f"[OK] Master data loaded. (Users: {len(users)}, Groups: {len(groups)})")

    # 2. Fetch Data
    print(f"[IN] Fetching records from Pleasanter (Site ID: {site_id})...")
    items = get_pleasanter_data(base_url, api_key, site_id, args.year, args.month)
    if not items:
        print("[!] No records found for the specified period.")
        return
    print(f"[OK] Fetched {len(items)} records!")

    # 3. Process Rows
    print(f"[*] Processing records and mapping columns...")
    excel_rows = []
    for item in items:
        def get_hash(key):
            val = item.get(key, {})
            if isinstance(val, str) and val.startswith('{'):
                try: return ast.literal_eval(val)
                except: return {}
            return val if isinstance(val, dict) else {}

        date_h = get_hash('DateHash')
        num_h = get_hash('NumHash')
        class_h = get_hash('ClassHash')
        
        creator_id = str(int(item.get('Creator', 0))) if item.get('Creator') else "0"
        owner_id = str(int(item.get('Owner', 0))) if item.get('Owner') else creator_id
        creator_name = user_map.get(creator_id, f"User({creator_id})")
        owner_name = user_map.get(owner_id, creator_name)
        
        def map_choice(col_name, val):
            if col_name in choices_mappings:
                return choices_mappings[col_name].get(str(val), val)
            return val

        row_data = {
            "ID": item.get('ResultId'),
            "提出日": date_h.get('DateA', '').split('T')[0] if date_h.get('DateA') else "",
            "氏名": owner_name,
            "部署名": group_map.get(str(class_h.get('ClassA', '')), ""),
            "職場": map_choice('ClassE', class_h.get('ClassE', '')),
            "件名": item.get('Title', ''),
            "区分": map_choice('ClassB', class_h.get('ClassB', '')),
            "効果金額・率(円・%)": num_h.get('NumA', 0),
            "効果率(%)適用": "適用" if item.get('CheckHash', {}).get('CheckA') else "",
            "改善効果点数": num_h.get('NumB', 0),
            "ヨコテン評価点数": num_h.get('NumC', 0),
            "一次審査日": date_h.get('DateB', '').split('T')[0] if date_h.get('DateB') else "",
            "一次審査者": user_map.get(str(class_h.get('ClassC', '')), ""),
            "創造力1": num_h.get('NumD', 0),
            "改善力1": num_h.get('NumE', 0),
            "問題解決力1": num_h.get('NumF', 0),
            "合計1": num_h.get('NumG', 0),
            "二次審査日": date_h.get('DateC', '').split('T')[0] if date_h.get('DateC') else "",
            "二次審査者": user_map.get(str(class_h.get('ClassD', '')), ""),
            "創造力2": num_h.get('NumH', 0),
            "改善力2": num_h.get('NumI', 0),
            "問題解決力2": num_h.get('NumJ', 0),
            "合計2": num_h.get('NumK', 0),
            "総合点数": num_h.get('NumL', 0),
            "内容": item.get('Body') if item.get('Body') else item.get('Title'),
            "受付状況": map_choice('Status', item.get('Status', '')),
            "ロック": "True" if item.get('Locked') else "False",
            "コメント": ""
        }
        excel_rows.append(row_data)

    df_full = pd.DataFrame(excel_rows)
    # Use hardcoded column order
    for c in FINAL_COLUMNS:
        if c not in df_full.columns: df_full[c] = ""
    df_full = df_full[FINAL_COLUMNS]
    df_high = df_full[df_full["総合点数"] >= 21].copy()

    # --- Sorting Logic ---
    print(f"[*] Sorting data for Excel sheets...")
    # Winners: Sorted by Total Score (Descending) as per notification text
    df_high = df_high.sort_values(by=["総合点数"], ascending=False).reset_index(drop=True)
    
    # All Records: Sorted by Dept, Workplace, and Total Score (Descending) as per notification text
    sort_keys_full = ["部署名", "職場", "総合点数"]
    df_full = df_full.sort_values(by=sort_keys_full, ascending=False).reset_index(drop=True)

    # 4. Save Excel
    print(f"[Excel] Saving report: {os.path.basename(excel_output)}...")
    try:
        with pd.ExcelWriter(excel_output, engine='openpyxl') as writer:
            df_high.to_excel(writer, sheet_name='高得点リスト(21点以上)', index=False)
            df_full.to_excel(writer, sheet_name='全件リスト', index=False)
        
        show_cols = ["氏名", "部署名", "職場", "件名", "区分", "効果金額・率(円・%)", "効果率(%)適用", "改善効果点数", "ヨコテン評価点数", "合計1", "合計2", "総合点数"]
        wb = openpyxl.load_workbook(excel_output)
        max_widths = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for col_idx, cell in enumerate(ws[1], 1):
                col_name = cell.value
                if col_name in show_cols:
                    current_max = max_widths.get(col_name, get_actual_width(col_name))
                    for row in ws.iter_rows(min_row=2):
                        val = row[col_idx-1].value
                        current_max = max(current_max, get_actual_width(val))
                    max_widths[col_name] = current_max

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for col_idx, cell in enumerate(ws[1], 1):
                col_letter = get_column_letter(col_idx)
                col_name = cell.value
                if col_name not in show_cols:
                    ws.column_dimensions[col_letter].hidden = True
                else:
                    ws.column_dimensions[col_letter].width = max_widths[col_name] + 2
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
        wb.save(excel_output)
        print(f"[OK] Excel report saved successfully!")

        # 5. Generate Shortened URLs and Message
        print(f"[*] Generating dynamic TinyURLs for notification...")
        url_high = generate_pleasanter_viewer_url(base_url, site_id, args.year, args.month, min_score=21)
        url_full = generate_pleasanter_viewer_url(base_url, site_id, args.year, args.month)
        
        short_url_high = shorten_url(url_high)
        short_url_full = shorten_url(url_full)
        
        winners_count = len(df_high)
        total_count = len(df_full)
        
        msg = f"""TO:各位

業務改善報告書{args.month}月度分の審査結果です。
紙で提出された分の件名,区分,審査結果をPleasanterに入力しました。
フィルター、並べ替えを行った表示のURLです。

{args.month}月提出分入賞者：{winners_count}件　総合得点 21以上,総合点数を降り順表示
{short_url_high}
（TinyURLのページにてクリックが必要）

{args.month}月提出分：{total_count}件 　部署名,職場,総合点数を降り順表示
{short_url_full}
（TinyURLのページにてクリックが必要）

以上、よろしくお願いいたします。"""
        
        with open(text_output, 'w', encoding='utf-8') as f:
            f.write(msg)
        print(f"[OK] Notification text saved: {os.path.basename(text_output)}")
        
        print(f"\n--- All operations completed successfully! ---\n")

    except PermissionError:
        print(f"[ERROR] Permission denied. Is the Excel file open?")
    except Exception as e:
        print(f"[ERROR] An unexpected error occurred: {e}")

if __name__ == "__main__":
    main()
